"""
OwnersApi - 当前Project名称;
test_common - 在创建文件的对话框中指定的文件名;
galaxy - 当前用户名;
2021/4/1 10:03 上午 
"""

import allure
import openpyxl as openpyxl
import pytest
import json

from api_page.common_page import CommonPage
from api_page.wework_utils import WeWorkUtils

path = '../testdata/daopu_api.xlsx'
title_arr = []
value_arr = []
config_dict = {}
ids_arr = []


def get_data():
    # excel对象
    wb = openpyxl.load_workbook(path)

    # 获取配置信息
    wsheet2 = wb['Sheet2']
    config_arr = list(wsheet2['B'])
    for i in range(len(config_arr)):
        cell = config_arr[i]
        if i == 0:
            config_dict['baseurl'] = cell.value
        elif i == 1:
            config_dict['is_level'] = cell.value
        elif i == 2:
            config_dict['appkey'] = cell.value
        elif i == 3:
            config_dict['appsercet'] = cell.value
        elif i == 4:
            config_dict['uuid'] = cell.value
        elif i == 5:
            string = cell.value
            is_module = string.split("、")
            config_dict['is_module'] = is_module
        elif i == 6:
            config_dict['token'] = cell.value
        elif i == 7:
            config_dict['uid'] = cell.value
        elif i == 8:
            config_dict['iostype'] = cell.value

    #接口数据
    wsheet1 = wb['Sheet1']
    for row in wsheet1.values:
        if '用例编号' in row:
            title_arr = row
        else:
            module_list = config_dict['is_module']
            if len(module_list) > 0:
                #只执行某些模块
                module = row[1]
                if module in module_list:
                    value_arr.append(row)
            else:
                value_arr.append(row)

    #排序
    if config_dict['is_level'] == 'Y':
        value_arr.sort(key=lambda x: x[4])

    #排序后的用例名
    for item_arr in value_arr:
        ids_arr.append(item_arr[2]+':'+item_arr[3])
    print('===========================')
    return value_arr

def write_excel_data(num,result,isSuccess):
    success = "不通过"
    if isSuccess == True:
        success = "通过"

    # excel对象
    wb = openpyxl.load_workbook(path)
    # 获取配置信息
    wsheet1 = wb['Sheet1']
    for col in wsheet1.iter_cols(min_col=1, max_col=1):
        for cell in col:
            if num == cell.value:
                row = cell.row
                print(row)
                wsheet1["L" + str(row)] = result
                wsheet1["M" + str(row)] = success
                wb.save(path)

class TestCommon():

    def setup(self):
        self.common = CommonPage()

    # @pytest.mark.skip()
    @pytest.mark.parametrize(
        "test_number,module,api_name,test_title,level,api_url,request_mode,condition,header,params,expect_result,actual_result,success,remark,assert_level",
        get_data(),ids=ids_arr)
    def test_common(self, test_number, module, api_name, test_title, level, api_url, request_mode, condition, header,
                    params, expect_result, actual_result, success, remark,assert_level):
        #请求参数
        if params is not None:
            params = json.loads(params)
        else:
            params = {}
        #请求头
        if header is not None:
            header = json.loads(header)
        else:
            header = {}

        #如果是密码登录 需要验签
        if api_name == '密码登录' and 'nil' not in header.values():
            t = WeWorkUtils.get_timestamp()
            t = str(t)
            password = WeWorkUtils.md5vale(params['password'] + config_dict['appsercet'])
            params['password'] = password
            sign = WeWorkUtils.md5vale(params['cellphone'] + password + config_dict['appkey'] + t + config_dict['appsercet'])
            header = {
                "appkey": config_dict['appkey'],
                "t": t,
                "uuid": config_dict['uuid'],
                "sign": sign,
            }
        elif condition == '登录':
            #需要登录前置条件
            t = WeWorkUtils.get_timestamp()
            t = str(t)
            sign = WeWorkUtils.md5vale(config_dict['appkey'] + config_dict['uuid'] + config_dict['token'] + t )
            header['appkey'] = config_dict['appkey']
            header['uuid'] = config_dict['uuid']
            header['uid'] = str(config_dict['uid'])
            header['token'] = config_dict['token']
            header['t'] = t
            header['iostype'] = str(config_dict['iostype'])
            header['sign'] = sign

        #url拼接
        url = config_dict['baseurl']+api_url
        #请求
        login_message = self.common.get_api(url=url,request_mode=request_mode, params=params, headers=header)
        #断言
        expect_result = json.loads(expect_result)
        isSuccess = False
        if assert_level == 'A' or assert_level == '':
            isSuccess = login_message['code'] == expect_result['code']
        elif assert_level == 'B':
            isSuccess = login_message['code'] == expect_result['code'] and login_message['msg'] == expect_result['msg']

        assert isSuccess

        result_str = json.dumps(login_message)
        write_excel_data(test_number,result_str,isSuccess)




if __name__=="__main__":

    pytest.main(["-s","test_common.py"])